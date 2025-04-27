import sys
import os
import re  # 添加正则表达式支持，用于公式解析
import json  # 导入json模块用于条件组的导入导出
import pandas as pd
import openpyxl
import shutil
import traceback
import importlib
import importlib.util
import subprocess  # 添加subprocess模块用于安装依赖
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QPushButton, QLineEdit, QListWidget, QFileDialog, 
    QMessageBox, QGroupBox, QCheckBox, QTabWidget, QProgressBar,
    QScrollArea, QFrame, QSizePolicy, QSplitter, QTableWidget,
    QTableWidgetItem, QListWidgetItem, QGridLayout
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from copy import copy

# 公式处理助手类
class FormulaHelper:
    @staticmethod
    def create_row_mapping(worksheet):
        """创建原始行号到行内容的映射"""
        row_content_map = {}
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), 1):
            # 存储该行的所有单元格值作为内容特征
            row_content = [str(cell.value) if cell.value is not None else "" for cell in row]
            row_content_map[row_idx] = row_content
        return row_content_map
    
    @staticmethod
    def build_row_mapping_after_deletion(original_map, deleted_rows):
        """在删除行后，建立原始行号到新行号的映射关系"""
        mapping = {}
        offset = 0
        
        for orig_row in sorted(original_map.keys()):
            if orig_row in deleted_rows:
                # 该行被删除，映射到None
                mapping[orig_row] = None
                offset += 1
            else:
                # 该行保留，新行号 = 原行号 - 偏移量
                mapping[orig_row] = orig_row - offset
        
        return mapping
    
    @staticmethod
    def parse_cell_references(formula):
        """解析公式中的单元格引用"""
        if not formula or not isinstance(formula, str):
            return []
        
        # 匹配Excel单元格引用的正则表达式模式
        # 匹配形式如A1, $A$1, Sheet1!A1, 'Sheet with spaces'!$A$1等
        pattern = r'(?:(?:[\'"]?[\w\s]+[\'"]?)?!)?(\$?[A-Z]+\$?[0-9]+)'
        
        # 查找所有匹配项
        matches = re.findall(pattern, formula)
        return matches
    
    @staticmethod
    def adjust_formula_references(formula, row_mapping):
        """根据行映射关系调整公式中的行引用"""
        if not formula or not isinstance(formula, str) or not formula.startswith('='):
            return formula
        
        # 解析公式中的单元格引用
        cell_refs = FormulaHelper.parse_cell_references(formula)
        
        # 按长度排序，确保替换最长的引用（避免部分替换导致错误）
        cell_refs.sort(key=len, reverse=True)
        
        new_formula = formula
        
        for ref in cell_refs:
            # 解析列引用和行引用
            match = re.match(r'(\$?[A-Z]+)(\$?[0-9]+)', ref)
            if match:
                col_ref, row_ref = match.groups()
                
                # 检查是否包含行号
                if row_ref.startswith('$'):
                    # 绝对引用，保留$符号
                    row_num = int(row_ref[1:])
                    is_absolute = True
                else:
                    # 相对引用
                    row_num = int(row_ref)
                    is_absolute = False
                
                # 如果行号在映射关系中
                if row_num in row_mapping:
                    new_row_num = row_mapping[row_num]
                    
                    # 如果映射到None，表示该行已被删除，保持原样（可能导致#REF!错误）
                    if new_row_num is None:
                        continue
                    
                    # 构建新的行引用
                    if is_absolute:
                        new_row_ref = f'${new_row_num}'
                    else:
                        new_row_ref = str(new_row_num)
                    
                    # 构建新的单元格引用
                    new_ref = f'{col_ref}{new_row_ref}'
                    
                    # 替换公式中的引用
                    # 需要考虑引用可能前后有特殊字符的情况
                    new_formula = new_formula.replace(ref, new_ref)
        
        return new_formula
    
    @staticmethod
    def update_formulas_in_sheet(worksheet, row_mapping):
        """更新工作表中所有单元格的公式"""
        updated_count = 0
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 如果单元格包含公式
                    original_formula = cell.value
                    if original_formula and isinstance(original_formula, str):
                        # 调整公式引用
                        new_formula = FormulaHelper.adjust_formula_references(original_formula, row_mapping)
                        
                        # 如果公式有变化，更新单元格
                        if new_formula != original_formula:
                            cell.value = new_formula
                            updated_count += 1
        
        return updated_count

# 最基本的依赖安装函数
def install_package(package):
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package],
                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        print(f"已安装 {package}")
        return True
    except Exception as e:
        print(f"安装 {package} 失败: {str(e)}")
        return False

# 首先检查setuptools是否安装
try:
    import pkg_resources
except ImportError:
    print("正在安装setuptools...")
    if not install_package("setuptools"):
        print("无法安装setuptools，程序无法继续。")
        sys.exit(1)
    try:
        import pkg_resources
    except ImportError:
        print("setuptools安装后仍无法导入pkg_resources，程序无法继续。")
        sys.exit(1)

# 检查单个包是否已安装
def is_package_installed(package_name):
    try:
        return importlib.util.find_spec(package_name) is not None
    except ModuleNotFoundError:
        return False

# 确保最小依赖项装载正确
minimal_deps = {
    'PyQt5': ['PyQt5.QtWidgets', 'PyQt5.QtCore'],
    'pandas': ['pandas'],
    'openpyxl': ['openpyxl']
}

# 初始化一个空的字典来存储未安装的依赖
missing_pkgs = {}

# 检查所有依赖
for pkg_name, modules in minimal_deps.items():
    missing = False
    for module in modules:
        if not is_package_installed(module):
            missing = True
            break
    
    if missing:
        missing_pkgs[pkg_name] = pkg_name

# 如果PyQt5未安装，需要先安装才能显示界面
if 'PyQt5' in missing_pkgs:
    print("正在安装PyQt5...")
    if install_package("PyQt5"):
        # 安装成功后从缺失列表中移除
        del missing_pkgs['PyQt5']
    else:
        print("错误: 无法安装PyQt5，程序无法继续。")
        print("请手动安装: pip install PyQt5")
        sys.exit(1)

# 现在可以导入PyQt的基本组件
from PyQt5.QtWidgets import (QApplication, QDialog, QVBoxLayout, QPushButton, 
                             QLabel, QProgressBar, QMessageBox, QTextEdit, QHBoxLayout,
                             QRadioButton, QButtonGroup, QStackedWidget, QScrollArea, QFrame)
from PyQt5.QtCore import Qt, QThread, pyqtSignal

# 在导入其余模块之前检查依赖
class DependencyInstaller(QThread):
    progress_signal = pyqtSignal(str, int)
    finished_signal = pyqtSignal(bool)
    
    def __init__(self, packages):
        super().__init__()
        self.packages = packages
        
    def run(self):
        total = len(self.packages)
        success = True
        
        for i, package in enumerate(self.packages):
            progress = int((i / total) * 100)
            self.progress_signal.emit(f"正在安装 {package}...", progress)
            
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package],
                                      stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                self.progress_signal.emit(f"{package} 安装成功", progress + 10)
            except:
                self.progress_signal.emit(f"{package} 安装失败", progress)
                success = False
                break
                
        self.finished_signal.emit(success)

class DependencyDialog(QDialog):
    """依赖安装对话框"""
    def __init__(self, missing_packages):
        super().__init__()
        self.missing_packages = missing_packages
        self.installer = None
        self.dependencies_installed = False
        self.initUI()
        
    def initUI(self):
        """初始化对话框界面"""
        self.setWindowTitle("依赖项安装")
        self.setMinimumWidth(500)
        self.setMinimumHeight(300)
        
        layout = QVBoxLayout()
        
        # 头部信息
        top_label = QLabel("需要安装以下依赖项才能运行此应用：")
        layout.addWidget(top_label)
        
        # 缺失依赖列表
        self.text_area = QTextEdit()
        self.text_area.setReadOnly(True)
        package_text = "\n".join([f"• {p}" for p in self.missing_packages])
        self.text_area.setText(package_text)
        layout.addWidget(self.text_area)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)
        
        # 状态标签
        self.status_label = QLabel("准备安装...")
        layout.addWidget(self.status_label)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        self.install_btn = QPushButton("安装依赖项")
        self.install_btn.clicked.connect(self.install_dependencies)
        btn_layout.addWidget(self.install_btn)
        
        self.manual_btn = QPushButton("手动安装说明")
        self.manual_btn.clicked.connect(self.show_manual_instructions)
        btn_layout.addWidget(self.manual_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        
    def install_dependencies(self):
        """开始安装依赖项"""
        self.install_btn.setEnabled(False)
        self.manual_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        
        # 在后台线程中安装依赖
        self.installer = DependencyInstaller(self.missing_packages)
        self.installer.progress_signal.connect(self.update_progress)
        self.installer.finished_signal.connect(self.installation_finished)
        self.installer.start()
    
    def update_progress(self, message, value):
        """更新安装进度"""
        self.status_label.setText(message)
        self.progress.setValue(value)
    
    def installation_finished(self, success):
        """安装完成处理"""
        if success:
            self.dependencies_installed = True
            QMessageBox.information(self, "安装完成", "所有依赖已成功安装！点击\"确定\"启动程序。")
            self.accept()
        else:
            self.status_label.setText("安装失败。请尝试手动安装。")
            self.manual_btn.setEnabled(True)
            self.cancel_btn.setEnabled(True)
    
    def show_manual_instructions(self):
        """显示手动安装说明"""
        instructions = "请在命令行中运行以下命令安装依赖：\n\n"
        for package in self.missing_packages:
            instructions += f"pip install {package}\n"
        
        self.text_area.setText(instructions)
        
    def closeEvent(self, event):
        """关闭事件处理"""
        if self.installer and self.installer.isRunning():
            self.installer.terminate()
        event.accept()

# 检查并安装必要的依赖
def check_dependencies():
    # 使用前面已经检测到的缺失依赖
    missing_packages = list(missing_pkgs.values())
    
    # 如果没有缺失的依赖，直接返回True
    if not missing_packages:
        return True
    
    # 初始化QApplication
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
        created_app = True
    else:
        created_app = False
    
    # 显示依赖对话框
    dialog = DependencyDialog(missing_packages)
    
    # 模拟DEMO_MODE的行为，确保两个文件的行为一致
    result = dialog.exec_()
    
    if result == QDialog.Accepted:
        # 安装成功，重启应用
        print("依赖安装成功，应用将继续运行...")
        if created_app:
            app.quit()
        # 由于我们在installation_finished中已经设置了dependencies_installed，不需要额外判断
        return True
    else:
        # 用户取消安装
        print("依赖安装被取消")
        return False

# 确保依赖已安装后再导入其他模块
if not check_dependencies():
    sys.exit(1)

# 现在可以安全地导入其余的模块
from collections import defaultdict  # 导入defaultdict用于存储跨工作表引用

# 批处理条件组类，用于存储多sheet并拆的条件
class ConditionGroup:
    def __init__(self, name=""):
        self.name = name if name else "未命名条件组"
        self.conditions = []  # [{sheet, column, values}]
    
    def add_condition(self, sheet, column, values):
        """添加一个筛选条件"""
        self.conditions.append({
            'sheet': sheet,
            'column': column,
            'values': values
        })
    
    def remove_condition(self, index):
        """删除指定索引的筛选条件"""
        if 0 <= index < len(self.conditions):
            del self.conditions[index]
    
    def clear_conditions(self):
        """清空所有筛选条件"""
        self.conditions = []

# 多sheet批处理界面组件
class BatchProcessingWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent  # 父窗口，用于访问其方法和属性
        self.excel_file = None  # 当前Excel文件路径
        self.df_dict = {}  # 存储所有sheet的DataFrame
        
        # 条件组列表
        self.condition_groups = []  # 存储ConditionGroup对象
        self.current_group_index = -1  # 当前选中的条件组索引
        
        # 批处理模式标志
        self.current_mode = 'batch'  # 批处理组件固定为batch模式
        
        # 初始化UI
        self.initUI()
    
    def initUI(self):
        """初始化批处理界面"""
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # 文件选择区域 - 固定显示
        file_layout = QHBoxLayout()
        
        self.select_file_btn = QPushButton('选择Excel文件')
        self.select_file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.select_file_btn)
        
        self.file_path_label = QLabel('未选择文件')
        self.file_path_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        file_layout.addWidget(self.file_path_label)
        
        main_layout.addLayout(file_layout)
        
        # 创建可调整大小的分割窗口
        splitter = QSplitter(Qt.Horizontal)
        splitter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(splitter, 1)
        
        # 左侧：条件组列表区域
        left_group = QGroupBox("条件组列表")
        left_group.setObjectName("主要拆分条件")
        left_layout = QVBoxLayout(left_group)
        left_layout.setContentsMargins(5, 10, 5, 5)  # 减小内边距
        
        # 创建左侧滚动区域
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QFrame.NoFrame)
        left_content = QWidget()
        left_content_layout = QVBoxLayout(left_content)
        left_content_layout.setContentsMargins(0, 0, 0, 0)
        
        # 条件组列表滚动区域
        group_list_scroll = QScrollArea()
        group_list_scroll.setWidgetResizable(True)
        group_list_scroll.setMinimumHeight(200)
        
        self.group_list = QListWidget()
        self.group_list.itemClicked.connect(self.group_selected)
        
        group_list_scroll.setWidget(self.group_list)
        left_content_layout.addWidget(group_list_scroll)
        
        # 条件组操作按钮
        group_btn_layout = QHBoxLayout()
        
        self.add_group_btn = QPushButton('添加条件组')
        self.add_group_btn.clicked.connect(self.add_condition_group)
        group_btn_layout.addWidget(self.add_group_btn)
        
        self.remove_group_btn = QPushButton('删除条件组')
        self.remove_group_btn.clicked.connect(self.remove_condition_group)
        self.remove_group_btn.setEnabled(False)
        group_btn_layout.addWidget(self.remove_group_btn)
        
        left_content_layout.addLayout(group_btn_layout)
        
        # 添加导入/导出条件组按钮
        io_btn_layout = QHBoxLayout()
        
        self.import_groups_btn = QPushButton('导入条件组')
        self.import_groups_btn.clicked.connect(self.import_condition_groups)
        io_btn_layout.addWidget(self.import_groups_btn)
        
        self.export_groups_btn = QPushButton('导出条件组')
        self.export_groups_btn.clicked.connect(self.export_condition_groups)
        self.export_groups_btn.setEnabled(False)
        io_btn_layout.addWidget(self.export_groups_btn)
        
        left_content_layout.addLayout(io_btn_layout)
        
        # 设置左侧滚动区域
        left_scroll.setWidget(left_content)
        left_layout.addWidget(left_scroll)
        
        # 添加左侧面板到分割器
        splitter.addWidget(left_group)
        
        # 右侧：条件编辑区域
        right_group = QGroupBox("条件编辑")
        right_layout = QVBoxLayout(right_group)
        
        # 创建右侧滚动区域
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setFrameShape(QFrame.NoFrame)
        right_content = QWidget()
        right_content_layout = QVBoxLayout(right_content)
        right_content_layout.setContentsMargins(0, 0, 0, 0)
        
        # 条件组名称编辑
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("条件组名称:"))
        self.group_name_edit = QLineEdit()
        self.group_name_edit.setPlaceholderText("输入条件组名称")
        self.group_name_edit.textChanged.connect(self.update_group_name)
        name_layout.addWidget(self.group_name_edit)
        right_content_layout.addLayout(name_layout)
        
        # 条件表格滚动区域
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setMinimumHeight(200)
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        
        self.condition_table = QTableWidget()
        self.condition_table.setColumnCount(4)  # 工作表、列、值、操作
        self.condition_table.setHorizontalHeaderLabels(["工作表", "列", "值", "操作"])
        self.condition_table.setSelectionBehavior(QTableWidget.SelectRows)
        table_layout.addWidget(self.condition_table)
        
        table_scroll.setWidget(table_widget)
        right_content_layout.addWidget(table_scroll)
        
        # 添加条件按钮
        self.add_condition_btn = QPushButton('添加筛选条件')
        self.add_condition_btn.clicked.connect(self.add_condition_dialog)
        self.add_condition_btn.setEnabled(False)
        right_content_layout.addWidget(self.add_condition_btn)
        
        # 设置右侧滚动区域
        right_scroll.setWidget(right_content)
        right_layout.addWidget(right_scroll)
        
        # 添加分割组件到主分割窗口
        splitter.addWidget(right_group)
        splitter.setSizes([300, 600])  # 设置初始大小比例
        
        # 底部区域：处理按钮和进度条 - 固定显示
        bottom_layout = QVBoxLayout()
        
        # 处理按钮
        self.process_btn = QPushButton('开始批量处理')
        self.process_btn.setMinimumHeight(30)
        self.process_btn.clicked.connect(self.start_batch_processing)
        self.process_btn.setEnabled(False)
        bottom_layout.addWidget(self.process_btn)
        
        # 进度条和状态
        progress_layout = QHBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel('')
        progress_layout.addWidget(self.status_label)
        
        bottom_layout.addLayout(progress_layout)
        
        main_layout.addLayout(bottom_layout)
    
    def select_file(self):
        """选择Excel文件"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        
        if not file_path:
            return
            
        try:
            self.status_label.setText("正在读取Excel文件...")
            self.progress_bar.setVisible(True)
            self.progress_bar.setMaximum(0)  # 不确定进度
            QApplication.processEvents()
            
            self.selected_file = file_path
            self.excel_file = file_path  # 确保self.excel_file被赋值
            self.file_path_label.setText(os.path.basename(file_path))
            
            # 使用pandas直接读取文件，不显示任何对话框
            try:
                # 不显示任何预览窗口和对话框，直接读取
                import pandas as pd
                import warnings
                
                # 关闭pandas警告
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    
                    # 读取Excel文件的所有sheet
                    excel_file = pd.ExcelFile(file_path)
                    self.sheet_names = excel_file.sheet_names
                    
                    # 根据文件大小选择加载方式
                    file_size = os.path.getsize(file_path) / (1024 * 1024)  # 转换为MB
                    
                    # 确保df_dict已初始化
                    self.df_dict = {}
                    
                    if file_size > 10:  # 如果文件大于10MB使用openpyxl
                        self.use_pandas = False
                        self.workbook = openpyxl.load_workbook(file_path, data_only=False)
                        self.sheet_names = self.workbook.sheetnames
                        self.data = {}  # 兼容性字典
                        
                        # 对于大文件，仍然需要为每个sheet创建一个DataFrame以便于条件选择
                        for sheet in self.sheet_names:
                            try:
                                # 只读取前1000行来提取列名和预览数据
                                temp_df = pd.read_excel(file_path, sheet_name=sheet, nrows=1000)
                                self.df_dict[sheet] = temp_df
                            except Exception as e:
                                print(f"读取工作表 {sheet} 数据时出错: {str(e)}")
                    else:
                        self.use_pandas = True
                        self.data = {}  # 存储所有sheet的DataFrame
                        # 读取所有sheet的数据
                        for sheet in self.sheet_names:
                            self.data[sheet] = pd.read_excel(file_path, sheet_name=sheet)
                            self.df_dict[sheet] = self.data[sheet]  # 兼容性
                
                self.status_label.setText("Excel文件已加载")
                
                # 批处理模式下启用添加条件组按钮
                self.add_group_btn.setEnabled(True)
                
            except Exception as e:
                error_details = traceback.format_exc()
                print(f"读取Excel文件出错: {str(e)}\n{error_details}")
                QMessageBox.critical(self, "错误", f"读取Excel文件出错: {str(e)}")
                self.status_label.setText("文件加载失败")
                
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"读取Excel文件出错: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"读取Excel文件出错: {str(e)}")
            self.status_label.setText("文件加载失败")
        finally:
            self.progress_bar.setVisible(False)
    
    def add_condition_group(self):
        """添加新的条件组"""
        # 创建新的条件组
        new_group = ConditionGroup(f"条件组 {len(self.condition_groups) + 1}")
        self.condition_groups.append(new_group)
        
        # 更新条件组列表
        self.update_group_list()
        
        # 选中新添加的条件组
        self.group_list.setCurrentRow(len(self.condition_groups) - 1)
        self.group_selected(self.group_list.currentItem())
        
        # 启用处理按钮
        self.process_btn.setEnabled(True)
    
    def remove_condition_group(self):
        """删除当前选中的条件组"""
        if self.current_group_index >= 0 and self.current_group_index < len(self.condition_groups):
            # 删除条件组
            del self.condition_groups[self.current_group_index]
            
            # 更新列表
            self.update_group_list()
            
            # 重置当前索引
            self.current_group_index = -1
            self.group_name_edit.setText("")
            self.condition_table.setRowCount(0)
            self.add_condition_btn.setEnabled(False)
            
            # 禁用删除按钮
            self.remove_group_btn.setEnabled(False)
            
            # 根据条件组数量启用/禁用处理按钮
            self.process_btn.setEnabled(len(self.condition_groups) > 0)
    
    def update_group_list(self):
        """更新条件组列表显示"""
        self.group_list.clear()
        
        for group in self.condition_groups:
            item = QListWidgetItem(group.name)
            self.group_list.addItem(item)
        
        # 根据是否有条件组来启用/禁用导出按钮
        self.export_groups_btn.setEnabled(len(self.condition_groups) > 0)
    
    def group_selected(self, item):
        """选中条件组列表中的项"""
        selected_row = self.group_list.currentRow()
        
        if selected_row >= 0 and selected_row < len(self.condition_groups):
            self.current_group_index = selected_row
            selected_group = self.condition_groups[selected_row]
            
            # 更新条件组名称
            self.group_name_edit.setText(selected_group.name)
            
            # 更新条件表格
            self.update_condition_table()
            
            # 启用添加条件和删除组按钮
            self.add_condition_btn.setEnabled(True)
            self.remove_group_btn.setEnabled(True)
    
    def update_group_name(self, text):
        """更新当前条件组的名称"""
        if self.current_group_index >= 0 and self.current_group_index < len(self.condition_groups):
            # 更新条件组名称
            self.condition_groups[self.current_group_index].name = text
            
            # 更新列表显示
            self.group_list.item(self.current_group_index).setText(text)
    
    def update_condition_table(self):
        """更新条件表格内容"""
        self.condition_table.setRowCount(0)  # 清空表格
        
        if self.current_group_index < 0 or self.current_group_index >= len(self.condition_groups):
            return
            
        conditions = self.condition_groups[self.current_group_index].conditions
        
        for i, condition in enumerate(conditions):
            self.condition_table.insertRow(i)
            
            # 工作表列
            sheet_item = QTableWidgetItem(condition['sheet'])
            self.condition_table.setItem(i, 0, sheet_item)
            
            # 列名列
            column_item = QTableWidgetItem(condition['column'])
            self.condition_table.setItem(i, 1, column_item)
            
            # 值列
            values = condition['values']
            if len(values) > 3:
                values_text = f"{values[0]}, {values[1]}, {values[2]}... 等{len(values)}个值"
            else:
                values_text = ", ".join(str(v) for v in values)
            values_item = QTableWidgetItem(values_text)
            self.condition_table.setItem(i, 2, values_item)
            
            # 操作列（删除按钮）
            delete_btn = QPushButton("删除")
            delete_btn.setProperty("row", i)  # 存储行索引
            delete_btn.clicked.connect(self.remove_condition_from_table)
            self.condition_table.setCellWidget(i, 3, delete_btn)
        
        # 调整列宽
        self.condition_table.resizeColumnsToContents()
    
    def remove_condition_from_table(self):
        """从表格中删除条件"""
        btn = self.sender()
        if btn:
            row = btn.property("row")
            
            if self.current_group_index >= 0 and self.current_group_index < len(self.condition_groups):
                # 删除条件
                self.condition_groups[self.current_group_index].remove_condition(row)
                
                # 更新表格
                self.update_condition_table()
    
    def add_condition_dialog(self):
        """打开添加筛选条件的对话框"""
        if not self.excel_file or self.current_group_index < 0:
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("添加筛选条件")
        dialog.setMinimumWidth(500)
        dialog.setMinimumHeight(400)
        
        layout = QVBoxLayout(dialog)
        
        # 工作表选择
        sheet_layout = QVBoxLayout()
        sheet_layout.addWidget(QLabel("选择工作表:"))
        
        sheet_list = QListWidget()
        for sheet_name in self.df_dict.keys():
            sheet_list.addItem(sheet_name)
        sheet_layout.addWidget(sheet_list)
        
        # 列选择
        column_layout = QVBoxLayout()
        column_layout.addWidget(QLabel("选择列:"))
        
        column_list = QListWidget()
        column_layout.addWidget(column_list)
        
        # 值选择
        values_layout = QVBoxLayout()
        values_layout.addWidget(QLabel("选择值:"))
        
        values_list = QListWidget()
        values_list.setSelectionMode(QListWidget.MultiSelection)
        values_layout.addWidget(values_list)
        
        # 连接信号
        def sheet_selected():
            column_list.clear()
            values_list.clear()
            
            selected_items = sheet_list.selectedItems()
            if selected_items:
                selected_sheet = selected_items[0].text()
                if selected_sheet in self.df_dict:
                    df = self.df_dict[selected_sheet]
                    for column in df.columns:
                        column_list.addItem(str(column))
        
        def column_selected():
            values_list.clear()
            
            selected_sheet_items = sheet_list.selectedItems()
            selected_column_items = column_list.selectedItems()
            
            if selected_sheet_items and selected_column_items:
                selected_sheet = selected_sheet_items[0].text()
                selected_column = selected_column_items[0].text()
                
                if selected_sheet in self.df_dict:
                    df = self.df_dict[selected_sheet]
                    if selected_column in df.columns:
                        unique_values = df[selected_column].dropna().unique().tolist()
                        for value in sorted(unique_values, key=str):
                            values_list.addItem(str(value))
        
        sheet_list.itemClicked.connect(sheet_selected)
        column_list.itemClicked.connect(column_selected)
        
        # 组合布局
        selection_layout = QHBoxLayout()
        selection_layout.addLayout(sheet_layout)
        selection_layout.addLayout(column_layout)
        selection_layout.addLayout(values_layout)
        
        layout.addLayout(selection_layout)
        
        # 按钮区域
        buttons_layout = QHBoxLayout()
        
        ok_button = QPushButton("添加")
        cancel_button = QPushButton("取消")
        
        buttons_layout.addWidget(ok_button)
        buttons_layout.addWidget(cancel_button)
        
        layout.addLayout(buttons_layout)
        
        # 连接按钮信号
        cancel_button.clicked.connect(dialog.reject)
        
        def add_condition():
            selected_sheet_items = sheet_list.selectedItems()
            selected_column_items = column_list.selectedItems()
            selected_values_items = values_list.selectedItems()
            
            if not selected_sheet_items or not selected_column_items or not selected_values_items:
                QMessageBox.warning(dialog, "警告", "请选择工作表、列和至少一个值")
                return
                
            selected_sheet = selected_sheet_items[0].text()
            selected_column = selected_column_items[0].text()
            selected_values = [item.text() for item in selected_values_items]
            
            # 添加条件
            self.condition_groups[self.current_group_index].add_condition(
                selected_sheet, selected_column, selected_values)
            
            # 更新表格
            self.update_condition_table()
            
            dialog.accept()
        
        ok_button.clicked.connect(add_condition)
        
        # 显示对话框
        dialog.exec_()
    
    def start_batch_processing(self):
        """开始批量处理"""
        if not self.excel_file or not self.condition_groups:
            QMessageBox.warning(self, "警告", "请先选择Excel文件并添加条件组")
            return
            
        # 确认对话框
        reply = QMessageBox.question(self, '确认', 
                                    f'将根据 {len(self.condition_groups)} 个条件组生成Excel文件，是否继续？',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        
        if reply == QMessageBox.No:
            return
            
        # 开始处理
        try:
            # 设置进度条
            self.progress_bar.setVisible(True)
            self.progress_bar.setMaximum(len(self.condition_groups))
            self.progress_bar.setValue(0)
            
            # 获取原始文件名和路径
            file_dir, file_name = os.path.split(self.excel_file)
            file_name_without_ext = os.path.splitext(file_name)[0]
            
            # 验证原始文件是否可以被openpyxl正常打开
            try:
                self.status_label.setText('正在验证Excel文件...')
                QApplication.processEvents()  # 让UI响应
                test_wb = openpyxl.load_workbook(self.excel_file)
                test_wb.close()
                use_openpyxl_method = True
            except Exception as e:
                error_msg = str(e)
                QMessageBox.warning(self, '警告', 
                    f'您的Excel文件包含一些不标准格式，将使用替代方法处理。\n'
                    f'某些复杂的格式可能无法完全保留。\n'
                    f'原因: {error_msg}')
                use_openpyxl_method = False
            
            # 处理每个条件组
            processed_files = []
            for i, group in enumerate(self.condition_groups):
                self.status_label.setText(f'正在处理条件组: {group.name} ({i+1}/{len(self.condition_groups)})')
                QApplication.processEvents()  # 让UI响应
                
                try:
                    # 安全的文件名
                    safe_name = "".join(c if c.isalnum() or c in [' ', '_', '-'] else '_' for c in group.name)
                    if len(safe_name) > 50:
                        safe_name = safe_name[:50]
                    
                    # 新文件路径
                    new_file_path = os.path.join(file_dir, f"{file_name_without_ext}_{safe_name}.xlsx")
                    
                    if use_openpyxl_method:
                        # 使用openpyxl方法处理
                        self.batch_process_with_openpyxl(group, new_file_path)
                    else:
                        # 使用pandas方法处理
                        self.batch_process_with_pandas(group, new_file_path)
                    
                    processed_files.append(os.path.basename(new_file_path))
                    
                except Exception as e:
                    error_details = traceback.format_exc()
                    print(f"处理条件组 '{group.name}' 时出错: {str(e)}\n{error_details}")
                    QMessageBox.warning(self, '警告', 
                                      f'处理条件组 "{group.name}" 时出错: {str(e)}\n'
                                      f'将跳过此条件组并继续处理其他组。')
                
                # 更新进度条
                self.progress_bar.setValue(i + 1)
                QApplication.processEvents()  # 让UI响应
            
            # 完成处理
            if processed_files:
                self.status_label.setText('批量处理完成!')
                result_message = f'已成功生成 {len(processed_files)} 个文件:\n\n'
                result_message += '\n'.join(processed_files[:10])
                if len(processed_files) > 10:
                    result_message += f'\n... 等总共 {len(processed_files)} 个文件'
                result_message += f'\n\n所有文件已保存在:\n{file_dir}'
                
                QMessageBox.information(self, '成功', result_message)
            else:
                QMessageBox.warning(self, '警告', '处理过程完成，但没有生成任何文件。')
            
        except Exception as e:
            error_details = traceback.format_exc()
            QMessageBox.critical(self, '错误', f'批量处理时出错: {str(e)}\n\n详细信息:\n{error_details}')
        finally:
            self.progress_bar.setVisible(False)
            self.status_label.setText('')
    
    def batch_process_with_openpyxl(self, condition_group, output_path):
        """使用openpyxl方法处理条件组"""
        try:
            # 复制原始文件
            shutil.copy2(self.excel_file, output_path)
            
            # 打开新文件
            wb = openpyxl.load_workbook(output_path, keep_vba=True, data_only=False, keep_links=True)
            
            # 存储所有工作表的行映射关系
            sheet_row_mappings = {}
            
            # 处理每个条件
            for condition in condition_group.conditions:
                sheet_name = condition['sheet']
                column_name = condition['column']
                filter_values = [str(v) for v in condition['values']]  # 转换所有值为字符串以便比较
                
                if sheet_name not in wb.sheetnames:
                    print(f"警告: 找不到工作表 '{sheet_name}'，跳过此筛选条件")
                    continue
                    
                # 获取工作表
                ws = wb[sheet_name]
                
                # 查找列索引
                header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
                col_index = -1
                for idx, cell in enumerate(header_row):
                    if str(cell.value) == column_name:
                        col_index = idx
                        break
                
                if col_index == -1:
                    print(f"警告: 在工作表 '{sheet_name}' 中找不到列 '{column_name}'，跳过此筛选条件")
                    continue
                
                # 创建行映射
                self.status_label.setText(f'处理工作表 {sheet_name} 的公式映射关系')
                QApplication.processEvents()
                original_row_map = FormulaHelper.create_row_mapping(ws)
                
                # 收集要删除的行
                rows_to_delete = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell_value = str(row[col_index].value) if row[col_index].value is not None else ""
                    if cell_value not in filter_values:
                        rows_to_delete.append(row_idx)
                
                # 如果所有行都要删除，保留一行数据避免工作表为空
                if len(rows_to_delete) >= ws.max_row - 1:
                    print(f"警告: 工作表 '{sheet_name}' 的筛选条件 '{column_name}' 将删除所有行，保留第一行数据")
                    if 2 in rows_to_delete:
                        rows_to_delete.remove(2)
                
                # 构建行映射关系
                row_mapping = FormulaHelper.build_row_mapping_after_deletion(original_row_map, rows_to_delete)
                sheet_row_mappings[sheet_name] = row_mapping
                
                # 删除不符合条件的行
                for row_idx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_idx, 1)
                
                # 更新公式
                FormulaHelper.update_formulas_in_sheet(ws, row_mapping)
            
            # 处理跨工作表公式引用
            self.status_label.setText(f'处理跨工作表公式引用...')
            QApplication.processEvents()
            
            # 遍历所有工作表，处理跨表引用
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                try:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.data_type == 'f':  # 如果单元格包含公式
                                formula = cell.value
                                if formula and isinstance(formula, str):
                                    new_formula = formula
                                    # 查找工作表引用，如"Sheet1!A1"
                                    sheet_refs = re.findall(r'(\'?[^!]+\'?)!(\$?[A-Z]+\$?[0-9]+)', formula)
                                    for ref_sheet, cell_ref in sheet_refs:
                                        # 清除引号
                                        clean_sheet_name = ref_sheet.strip("'")
                                        # 如果引用的工作表有行映射
                                        if clean_sheet_name in sheet_row_mappings:
                                            # 解析引用
                                            match = re.match(r'(\$?[A-Z]+)(\$?[0-9]+)', cell_ref)
                                            if match:
                                                col_ref, row_ref = match.groups()
                                                
                                                # 获取行号
                                                if row_ref.startswith('$'):
                                                    row_num = int(row_ref[1:])
                                                    is_absolute = True
                                                else:
                                                    row_num = int(row_ref)
                                                    is_absolute = False
                                                
                                                # 查找新行号
                                                mapping = sheet_row_mappings[clean_sheet_name]
                                                if row_num in mapping:
                                                    new_row_num = mapping[row_num]
                                                    
                                                    if new_row_num is None:
                                                        continue
                                                    
                                                    # 构建新引用
                                                    if is_absolute:
                                                        new_row_ref = f'${new_row_num}'
                                                    else:
                                                        new_row_ref = str(new_row_num)
                                                    
                                                    new_cell_ref = f'{col_ref}{new_row_ref}'
                                                    
                                                    # 替换引用
                                                    original_ref = f'{ref_sheet}!{cell_ref}'
                                                    replacement = f'{ref_sheet}!{new_cell_ref}'
                                                    new_formula = new_formula.replace(original_ref, replacement)
                                    
                                    # 更新公式
                                    if new_formula != formula:
                                        try:
                                            cell.value = new_formula
                                        except Exception as e:
                                            print(f"更新公式时出错: {e} - 原始公式: {formula}, 新公式: {new_formula}")
                except Exception as e:
                    print(f"处理工作表 '{sheet_name}' 的公式时出错: {str(e)}")
            
            # 应用所有工作表的格式调整
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    # 处理合并单元格
                    if hasattr(ws, 'merged_cells') and ws.merged_cells:
                        print(f"处理工作表 {sheet_name} 的合并单元格")
                except Exception as e:
                    print(f"处理工作表 '{sheet_name}' 的合并单元格时出错: {str(e)}")
            
            # 保存文件
            try:
                wb.save(output_path)
                wb.close()
            except PermissionError:
                QMessageBox.critical(self, "错误", f"无法保存文件，可能是权限不足或文件被其他程序占用:\n{output_path}")
                raise
            
        except PermissionError:
            QMessageBox.critical(self, "错误", f"无法写入文件，可能是权限不足或文件被其他程序占用:\n{output_path}")
            raise
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"处理条件组时出错: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"处理条件组时出错: {str(e)}")
            raise
    
    def batch_process_with_pandas(self, condition_group, output_path):
        """使用pandas方法处理条件组"""
        try:
            # 显示警告
            QMessageBox.warning(self, '警告', 
                f'由于Excel文件格式问题，将使用备用方法处理条件组 "{condition_group.name}"。\n'
                f'注意：此方法将无法保留公式及其引用关系，只会保留当前计算结果。')
            
            # 读取所有工作表
            all_dfs = {}
            processed_sheets = []
            skipped_sheets = []
            
            for sheet_name, df in self.df_dict.items():
                try:
                    # 创建副本以防止修改原始数据
                    sheet_df = df.copy()
                    
                    # 应用条件组中的所有条件
                    filters_applied = False
                    for condition in condition_group.conditions:
                        if condition['sheet'] == sheet_name:
                            column = condition['column']
                            values = [str(v) for v in condition['values']]  # 转换为字符串进行比较
                            
                            # 应用筛选
                            if column in sheet_df.columns:
                                # 将数据列转换为字符串以进行比较
                                sheet_df = sheet_df[sheet_df[column].astype(str).isin(values)]
                                filters_applied = True
                                print(f"已为工作表 '{sheet_name}' 应用列 '{column}' 的筛选条件，筛选后行数: {len(sheet_df)}")
                            else:
                                print(f"警告: 在工作表 '{sheet_name}' 中找不到列 '{column}'，跳过此筛选条件")
                    
                    # 添加到结果字典
                    all_dfs[sheet_name] = sheet_df
                    if filters_applied:
                        processed_sheets.append(sheet_name)
                    else:
                        print(f"工作表 '{sheet_name}' 没有应用任何筛选条件，保留所有行")
                        
                except Exception as e:
                    print(f"处理工作表 '{sheet_name}' 时出错: {str(e)}")
                    skipped_sheets.append(sheet_name)
            
            # 如果没有成功处理任何工作表，则报告错误
            if not all_dfs:
                raise Exception("无法处理任何工作表")
            
            # 将所有工作表写入新文件
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet_name, sheet_df in all_dfs.items():
                        # 如果筛选后的数据为空，添加一个标题行
                        if len(sheet_df) == 0 and len(sheet_df.columns) > 0:
                            print(f"警告: 工作表 '{sheet_name}' 筛选后没有数据，只保留标题行")
                            # 创建一个只有标题的空DataFrame
                            empty_df = pd.DataFrame(columns=sheet_df.columns)
                            empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        else:
                            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                raise Exception(f"写入Excel文件时出错: {str(e)}")
                    
            # 生成处理报告
            report_message = f"已使用备用方法处理文件 \"{os.path.basename(output_path)}\"。\n\n"
            
            if processed_sheets:
                report_message += f"已应用筛选条件的工作表 ({len(processed_sheets)}):\n"
                report_message += ", ".join(processed_sheets[:5])
                if len(processed_sheets) > 5:
                    report_message += f"... 等 {len(processed_sheets)} 个工作表\n\n"
                else:
                    report_message += "\n\n"
            
            if skipped_sheets:
                report_message += f"处理失败的工作表 ({len(skipped_sheets)}):\n"
                report_message += ", ".join(skipped_sheets)
                report_message += "\n\n"
            
            report_message += "注意：某些复杂的格式可能未能完全保留。"
            
            # 提示用户可能丢失一些格式
            QMessageBox.information(self, '信息', report_message)
            
        except PermissionError:
            QMessageBox.critical(self, "错误", f"无法写入文件，可能是权限不足或文件被其他程序占用:\n{output_path}")
            raise
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"使用pandas处理条件组时出错: {str(e)}\n{error_details}")
            QMessageBox.critical(self, "错误", f"使用pandas处理条件组时出错: {str(e)}")
            raise
    
    def import_condition_groups(self):
        """导入条件组"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, '导入条件组', '', 'JSON文件 (*.json)')
        
        if not file_path:
            return
            
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError as e:
                    QMessageBox.critical(self, '错误', f'无效的JSON文件格式: {str(e)}')
                    return
                
            if not isinstance(data, list):
                QMessageBox.warning(self, '警告', '无效的条件组文件格式，应为条件组列表')
                return
                
            imported_groups = []
            skipped_groups = 0
            
            for group_data in data:
                if 'name' in group_data and 'conditions' in group_data:
                    group = ConditionGroup(group_data['name'])
                    valid_conditions = 0
                    
                    for cond in group_data['conditions']:
                        if all(k in cond for k in ['sheet', 'column', 'values']):
                            group.add_condition(cond['sheet'], cond['column'], cond['values'])
                            valid_conditions += 1
                        else:
                            print(f"跳过无效的条件: {cond}")
                    
                    if valid_conditions > 0:
                        imported_groups.append(group)
                    else:
                        print(f"跳过没有有效条件的组: {group_data['name']}")
                        skipped_groups += 1
                else:
                    skipped_groups += 1
                    print(f"跳过无效的条件组数据: {group_data}")
            
            if not imported_groups:
                QMessageBox.warning(self, '警告', '未发现有效的条件组')
                return
                
            # 询问是否替换现有条件组
            if self.condition_groups:
                reply = QMessageBox.question(self, '确认', 
                                          '是否替换现有条件组？选择"否"将添加到现有条件组',
                                          QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, 
                                          QMessageBox.No)
                
                if reply == QMessageBox.Cancel:
                    return
                elif reply == QMessageBox.Yes:
                    self.condition_groups = imported_groups
                else:  # No - 添加到现有条件组
                    self.condition_groups.extend(imported_groups)
            else:
                self.condition_groups = imported_groups
            
            # 更新UI
            self.update_group_list()
            if self.group_list.count() > 0:
                self.group_list.setCurrentRow(0)
                self.group_selected(self.group_list.item(0))
            
            # 启用处理按钮
            self.process_btn.setEnabled(True)
            
            # 提示信息
            success_msg = f'已导入 {len(imported_groups)} 个条件组'
            if skipped_groups > 0:
                success_msg += f'，跳过了 {skipped_groups} 个无效条件组'
            QMessageBox.information(self, '成功', success_msg)
            
        except FileNotFoundError:
            QMessageBox.critical(self, '错误', f'找不到文件: {file_path}')
        except PermissionError:
            QMessageBox.critical(self, '错误', f'无法读取文件，可能是权限不足: {file_path}')
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"导入条件组时出错: {str(e)}\n{error_details}")
            QMessageBox.critical(self, '错误', f'导入条件组时出错: {str(e)}\n\n详细信息:\n{error_details}')
    
    def export_condition_groups(self):
        """导出条件组"""
        if not self.condition_groups:
            QMessageBox.warning(self, '警告', '没有条件组可以导出')
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, '导出条件组', '', 'JSON文件 (*.json)')
        
        if not file_path:
            return
        
        # 确保文件扩展名为.json
        if not file_path.lower().endswith('.json'):
            file_path += '.json'
            
        try:
            # 将条件组转换为可序列化的字典
            groups_data = []
            for group in self.condition_groups:
                # 确保条件组有有效的名称和条件
                if not hasattr(group, 'name') or not hasattr(group, 'conditions'):
                    print(f"跳过无效的条件组: {group}")
                    continue
                    
                group_data = {
                    'name': group.name,
                    'conditions': group.conditions
                }
                groups_data.append(group_data)
            
            if not groups_data:
                QMessageBox.warning(self, '警告', '没有有效的条件组可以导出')
                return
                
            # 写入JSON文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(groups_data, f, ensure_ascii=False, indent=2)
                
            QMessageBox.information(self, '成功', f'已成功导出 {len(groups_data)} 个条件组到文件:\n{file_path}')
            
        except PermissionError:
            QMessageBox.critical(self, '错误', f'无法写入文件，可能是权限不足或文件被其他程序占用:\n{file_path}')
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"导出条件组时出错: {str(e)}\n{error_details}")
            QMessageBox.critical(self, '错误', f'导出条件组时出错: {str(e)}\n\n详细信息:\n{error_details}')

    def copy_cell_format(self, source_cell, target_cell):
        """完整复制单元格格式"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    def copy_sheet_formatting(self, source_sheet, target_sheet):
        """复制工作表级别的格式设置"""
        # 复制列宽
        for column_letter, column_dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[column_letter].width = column_dim.width
        
        # 复制行高
        for row_number, row_dim in source_sheet.row_dimensions.items():
            if row_dim.height is not None:
                target_sheet.row_dimensions[row_number].height = row_dim.height

class ExcelSplitterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        # 初始化主窗口
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Excel拆分工具 - 多sheet并拆')
        self.setGeometry(100, 100, 950, 600)  # 设置适合Mac Dock栏的窗口高度
        
        # 创建主窗口部件和布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # 直接创建并添加批处理模式控件
        self.batch_mode_widget = BatchProcessingWidget(self)
        main_layout.addWidget(self.batch_mode_widget)
        
        # 将窗口居中显示
        self.center()
    
    def center(self):
        """将窗口居中显示在屏幕上"""
        screen = QApplication.desktop().screenGeometry()
        size = self.geometry()
        self.move((screen.width() - size.width()) // 2, 
                 (screen.height() - size.height()) // 2)

def main():
    app = QApplication(sys.argv)
    window = ExcelSplitterApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main() 